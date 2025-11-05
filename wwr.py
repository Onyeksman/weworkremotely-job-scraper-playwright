from playwright.sync_api import sync_playwright
import json
import time
import csv
import pandas as pd
from urllib.parse import urljoin
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def scrape_wwr_professional():
    """
    Professional WeWorkRemotely Scraper
    - Scrapes in exact sequential order
    - Exports to beautifully formatted Excel & CSV
    - Clean, professional output ready for clients
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,  # Set to True for background operation
            args=['--disable-blink-features=AutomationControlled']
        )

        context = browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )

        page = context.new_page()
        all_jobs = []

        print("=" * 80)
        print("🔍 PROFESSIONAL WeWorkRemotely SCRAPER")
        print("=" * 80)
        print(f"⏰ Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        try:
            # Load main page with retry logic
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    print(f"Loading main page (attempt {attempt + 1}/{max_retries})...")
                    page.goto('https://weworkremotely.com/', timeout=60000, wait_until='load')
                    page.wait_for_selector('li.feature', state='visible', timeout=30000)
                    print("✅ Main page loaded successfully!\n")
                    break
                except Exception as e:
                    if attempt == max_retries - 1:
                        raise
                    print(f"⚠️  Retry due to: {e}")
                    time.sleep(2)

            # Get all job cards IN EXACT ORDER
            job_cards = page.query_selector_all('li.feature')
            total_jobs = len(job_cards)

            print(f"📊 Found {total_jobs} total jobs")
            print(f"🎯 Scraping in exact sequential order as displayed on website\n")
            print("-" * 80 + "\n")

            # Scrape each job IN ORDER
            for index, card in enumerate(job_cards, 1):
                try:
                    job_data = {}

                    # ============================================================
                    # STEP 1: Extract data from listing card
                    # ============================================================

                    # Job URL
                    link_elem = card.query_selector('a[href^="/remote-jobs/"]')
                    if not link_elem:
                        print(f"[{index}/{total_jobs}] ⏭️  Skipping - no valid link found\n")
                        continue

                    # Check if listing is locked (requires login)
                    class_attr = link_elem.get_attribute('class') or ''
                    is_locked = 'listing-link--locked' in class_attr

                    href = link_elem.get_attribute('href')
                    job_data['Job URL'] = urljoin('https://weworkremotely.com', href)

                    # Job Title
                    title_elem = card.query_selector('.new-listing__header__title')
                    job_data['Job Title'] = title_elem.inner_text().strip() if title_elem else 'N/A'

                    # Company Name
                    company_elem = card.query_selector('.new-listing__company-name')
                    company_text = company_elem.inner_text().strip() if company_elem else 'N/A'
                    job_data['Company Name'] = company_text

                    # Company Location/Headquarters
                    location_elem = card.query_selector('.new-listing__company-headquarters')
                    job_data['Company Headquarters'] = location_elem.inner_text().strip() if location_elem else 'N/A'

                    # Company Logo URL
                    logo_elem = card.query_selector('.tooltip--flag-logo__flag-logo')
                    if logo_elem:
                        style = logo_elem.get_attribute('style')
                        if style and 'url(' in style:
                            logo_url = style.split('url(')[1].split(')')[0]
                            job_data['Company Logo URL'] = logo_url
                        else:
                            job_data['Company Logo URL'] = 'N/A'
                    else:
                        job_data['Company Logo URL'] = 'N/A'

                    # Company Profile URL
                    company_link_elem = card.query_selector('a[href^="/company/"]')
                    if company_link_elem:
                        company_href = company_link_elem.get_attribute('href')
                        job_data['Company Profile URL'] = urljoin('https://weworkremotely.com', company_href)
                    else:
                        job_data['Company Profile URL'] = 'N/A'

                    # Categories/Tags (Featured, Top 100, Full-Time, etc.)
                    categories = []
                    category_elems = card.query_selector_all('.new-listing__categories__category')
                    for cat in category_elems:
                        cat_text = cat.inner_text().strip()
                        categories.append(cat_text)

                    job_data['Tags'] = ', '.join(categories) if categories else 'N/A'

                    print(f"[{index}/{total_jobs}] 📝 {job_data['Job Title']}")
                    print(f"              🏢 {job_data['Company Name']}")

                    # ============================================================
                    # STEP 2: Scrape detail page (if not locked)
                    # ============================================================

                    if is_locked:
                        print(f"              🔒 Locked listing - basic info only\n")
                        # Set default values for locked listings
                        job_data['Date Posted'] = 'N/A'
                        job_data['Apply Deadline'] = 'N/A'
                        job_data['Job Type'] = 'N/A'
                        job_data['Job Category'] = 'N/A'
                        job_data['Region'] = 'N/A'
                        job_data['Salary'] = 'N/A'
                        job_data['Company Description'] = 'N/A'
                        job_data['Job Description'] = 'N/A'
                        job_data['Application URL'] = 'N/A'
                        job_data['Company Total Jobs Posted'] = 'N/A'

                        all_jobs.append(job_data)
                        continue

                    print(f"              📄 Fetching full details...")

                    # Open detail page
                    detail_page = context.new_page()

                    # Load with retry
                    for attempt in range(3):
                        try:
                            detail_page.goto(job_data['Job URL'], timeout=60000, wait_until='load')
                            detail_page.wait_for_selector('.lis-container__header__hero__company-info', timeout=20000)
                            break
                        except Exception as e:
                            if attempt == 2:
                                raise
                            time.sleep(1)

                    # Date Posted
                    posted_elem = detail_page.query_selector(
                        '.lis-container__header__hero__company-info__icons__item span')
                    job_data['Date Posted'] = posted_elem.inner_text().strip() if posted_elem else 'N/A'

                    # Company Description
                    company_desc_elem = detail_page.query_selector(
                        '.lis-container__header__hero__company-info__description')
                    job_data[
                        'Company Description'] = company_desc_elem.inner_text().strip() if company_desc_elem else 'N/A'

                    # Job Description (Full Text)
                    job_desc_elem = detail_page.query_selector('.lis-container__job__content__description')
                    if job_desc_elem:
                        # Get clean text version
                        job_desc_text = job_desc_elem.inner_text().strip()
                        job_data['Job Description'] = job_desc_text

                        # Extract salary if mentioned
                        salary_match = re.search(
                            r'\$[\d,]+(?:\s*-\s*\$[\d,]+)?(?:\s*(?:per|/)\s*(?:year|hour|month|annum|annually))?|€[\d,]+(?:\s*-\s*€[\d,]+)?(?:\s*(?:per|/)\s*(?:year|hour|month|annum|annually))?|£[\d,]+(?:\s*-\s*£[\d,]+)?',
                            job_desc_text, re.IGNORECASE)
                        job_data['Salary'] = salary_match.group(0) if salary_match else 'N/A'
                    else:
                        job_data['Job Description'] = 'N/A'
                        job_data['Salary'] = 'N/A'

                    # Application URL (Apply Button)
                    apply_btn = detail_page.query_selector('a.apply-btn:not(.apply-btn--locked)')
                    if apply_btn:
                        apply_href = apply_btn.get_attribute('href')
                        if apply_href and 'register' not in apply_href:
                            job_data['Application URL'] = urljoin('https://weworkremotely.com', apply_href)
                        else:
                            # Try to find external apply link in description
                            external_link = detail_page.query_selector(
                                '.lis-container__job__content__description a[href*="apply"], .lis-container__job__content__description a[href*="jobs"], .lis-container__job__content__description a[href*="careers"]')
                            if external_link:
                                job_data['Application URL'] = external_link.get_attribute('href')
                            else:
                                job_data['Application URL'] = 'See Job Description'
                    else:
                        job_data['Application URL'] = 'See Job Description'

                    # Sidebar Details
                    sidebar_items = detail_page.query_selector_all(
                        '.lis-container__job__sidebar__job-about__list__item')

                    # Initialize fields
                    job_data['Apply Deadline'] = 'N/A'
                    job_data['Job Type'] = 'N/A'
                    job_data['Job Category'] = 'N/A'
                    job_data['Region'] = 'N/A'

                    for item in sidebar_items:
                        text = item.inner_text().strip()

                        if 'Apply before' in text:
                            span_elem = item.query_selector('span')
                            job_data['Apply Deadline'] = span_elem.inner_text().strip() if span_elem else 'N/A'

                        elif 'Job type' in text:
                            job_type_elem = item.query_selector('.box--jobType')
                            job_data['Job Type'] = job_type_elem.inner_text().strip() if job_type_elem else 'N/A'

                        elif 'Category' in text:
                            category_elem = item.query_selector('.box--blue')
                            job_data['Job Category'] = category_elem.inner_text().strip() if category_elem else 'N/A'

                        elif 'Region' in text:
                            region_elem = item.query_selector('.box--region')
                            job_data['Region'] = region_elem.inner_text().strip() if region_elem else 'N/A'

                    # Company Total Jobs Posted
                    jobs_posted_elem = detail_page.query_selector(
                        '.lis-container__job__sidebar__companyDetails__info__jobs-posted')
                    if jobs_posted_elem:
                        jobs_posted_text = jobs_posted_elem.inner_text().strip()
                        # Extract number from "Jobs posted: 174"
                        match = re.search(r'\d+', jobs_posted_text)
                        job_data['Company Total Jobs Posted'] = match.group(0) if match else 'N/A'
                    else:
                        job_data['Company Total Jobs Posted'] = 'N/A'

                    detail_page.close()

                    all_jobs.append(job_data)
                    print(f"              ✅ Complete ({len(all_jobs)} scraped so far)\n")

                    # Be respectful - add delay
                    time.sleep(1.5)

                except Exception as e:
                    print(f"              ❌ Error: {str(e)}\n")
                    continue

        except Exception as e:
            print(f"\n❌ Fatal error: {str(e)}")

        finally:
            browser.close()

        # ============================================================
        # STEP 3: Export to PROFESSIONALLY FORMATTED files
        # ============================================================

        if all_jobs:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            print("\n" + "=" * 80)
            print("💾 CREATING PROFESSIONAL EXPORTS")
            print("=" * 80 + "\n")

            # Define professional column order
            column_order = [
                'Job Title',
                'Company Name',
                'Company Headquarters',
                'Date Posted',
                'Apply Deadline',
                'Job Type',
                'Job Category',
                'Region',
                'Salary',
                'Tags',
                'Company Description',
                'Job Description',
                'Job URL',
                'Application URL',
                'Company Profile URL',
                'Company Logo URL',
                'Company Total Jobs Posted'
            ]

            # Create DataFrame
            df = pd.DataFrame(all_jobs)

            # Reorder columns (only include columns that exist)
            existing_columns = [col for col in column_order if col in df.columns]
            df = df[existing_columns]

            # ============================================================
            # 1. PROFESSIONAL EXCEL with FORMATTING
            # ============================================================

            excel_filename = f'WeWorkRemotely_Jobs_{timestamp}.xlsx'

            # Save initial Excel file
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Remote Jobs')

            # Load workbook for professional formatting
            wb = load_workbook(excel_filename)
            ws = wb['Remote Jobs']

            # Define professional color scheme
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78',
                                      fill_type='solid')  # Professional dark blue
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # White bold text

            # Cell styling
            cell_font = Font(name='Calibri', size=10)
            cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Border style
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )

            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Set row height for header
            ws.row_dimensions[1].height = 25

            # Format data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # Adjust column widths professionally
            column_widths = {
                'Job Title': 35,
                'Company Name': 25,
                'Company Headquarters': 20,
                'Date Posted': 15,
                'Apply Deadline': 15,
                'Job Type': 12,
                'Job Category': 20,
                'Region': 25,
                'Salary': 20,
                'Tags': 30,
                'Company Description': 50,
                'Job Description': 60,
                'Job URL': 40,
                'Application URL': 40,
                'Company Profile URL': 35,
                'Company Logo URL': 35,
                'Company Total Jobs Posted': 15
            }

            for idx, column in enumerate(existing_columns, 1):
                column_letter = get_column_letter(idx)
                ws.column_dimensions[column_letter].width = column_widths.get(column, 15)

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Add auto-filter
            ws.auto_filter.ref = ws.dimensions

            # Alternate row coloring for better readability
            light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            for row_idx in range(3, ws.max_row + 1, 2):  # Every other row starting from row 3
                for cell in ws[row_idx]:
                    if cell.fill.start_color.rgb != header_fill.start_color.rgb:
                        cell.fill = light_gray_fill

            # Save formatted Excel
            wb.save(excel_filename)
            print(f"✅ Professional Excel saved: {excel_filename}")
            print(f"   • Header: Dark blue with white bold text")
            print(f"   • Borders: Clean professional borders")
            print(f"   • Frozen header row for easy scrolling")
            print(f"   • Auto-filter enabled")
            print(f"   • Alternating row colors")
            print(f"   • Optimized column widths")

            # ============================================================
            # 2. PROFESSIONAL CSV
            # ============================================================

            csv_filename = f'WeWorkRemotely_Jobs_{timestamp}.csv'
            df.to_csv(csv_filename, index=False, encoding='utf-8-sig')  # UTF-8 with BOM for Excel compatibility
            print(f"\n✅ Professional CSV saved: {csv_filename}")
            print(f"   • UTF-8 encoded with BOM (Excel compatible)")
            print(f"   • Clean, consistent formatting")

            # ============================================================
            # 3. CLEAN JSON (for backup/API use)
            # ============================================================

            json_filename = f'WeWorkRemotely_Jobs_{timestamp}.json'
            with open(json_filename, 'w', encoding='utf-8') as f:
                json.dump(all_jobs, f, indent=2, ensure_ascii=False)
            print(f"\n✅ JSON saved: {json_filename}")

            # ============================================================
            # 4. Summary Statistics
            # ============================================================

            print("\n" + "=" * 80)
            print("📊 SCRAPING SUMMARY")
            print("=" * 80)
            print(f"Total jobs found: {total_jobs}")
            print(f"Successfully scraped: {len(all_jobs)}")
            print(f"Success rate: {len(all_jobs) / total_jobs * 100:.1f}%")

            # Job types breakdown
            job_types = df['Job Type'].value_counts().to_dict()
            print(f"\nJob Types:")
            for jtype, count in job_types.items():
                print(f"  • {jtype}: {count}")

            # Job categories breakdown
            categories = df['Job Category'].value_counts().head(5).to_dict()
            print(f"\nTop 5 Categories:")
            for cat, count in categories.items():
                print(f"  • {cat}: {count}")

            # Regions breakdown
            regions = df['Region'].value_counts().head(5).to_dict()
            print(f"\nTop 5 Regions:")
            for region, count in regions.items():
                print(f"  • {region}: {count}")

            print(f"\n⏰ Finished at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print("=" * 80)
            print("\n🎉 SCRAPING COMPLETE - Professional files ready for presentation!")
            print("=" * 80)

        else:
            print("\n❌ No jobs were scraped successfully")

        return all_jobs


if __name__ == "__main__":
    jobs = scrape_wwr_professional()